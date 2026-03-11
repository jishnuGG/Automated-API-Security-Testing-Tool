"""
Export endpoint for downloading security logs as CSV, styled Excel, or PDF report.

GET /api/v1/export-logs?start_date=YYYY-MM-DD&end_date=YYYY-MM-DD&format=csv|xlsx|pdf

Requires JWT authentication. Exports only the authenticated user's logs.
"""

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from app.database import get_database
from app.dependencies.auth_dependency import get_current_user
from datetime import datetime, timedelta
import pandas as pd
import io
import sys

router = APIRouter()

# Columns to export and their display names
EXPORT_COLUMNS = {
    "domain": "Domain",
    "timestamp": "Timestamp",
    "method": "Method",
    "url": "URL",
    "status_code": "Status Code",
    "ml_probability": "ML Probability",
    "heuristic_score": "Heuristic Score",
    "risk_score": "Risk Score",
    "risk_level": "Risk Level",
    "threat_type": "Threat Type",
    "threat_label": "Threat Label",
    "owasp_category": "OWASP Category",
    "is_https": "Is HTTPS",
    "reasons": "Reasons",
}


async def _fetch_logs_for_export(
    user_id: str, start_date: str, end_date: str
) -> list[dict]:
    """Query high_risk_logs for the user within the date range."""
    db = await get_database()
    if db is None:
        raise HTTPException(status_code=503, detail="Database not available")

    try:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
    except ValueError:
        raise HTTPException(
            status_code=400,
            detail="Invalid date format. Use YYYY-MM-DD.",
        )

    query = {
        "user_id": user_id,
        "timestamp": {
            "$gte": start_dt,
            "$lt": end_dt
        },
    }

    cursor = db.high_risk_logs.find(query).sort("timestamp", -1)
    logs = await cursor.to_list(length=10000)

    return logs


def _logs_to_dataframe(logs: list[dict]) -> pd.DataFrame:
    """Convert raw MongoDB log dicts into a clean DataFrame."""
    rows = []
    for log in logs:
        row = {}
        for key, display in EXPORT_COLUMNS.items():
            val = log.get(key, "")
            if key == "timestamp" and val:
                try:
                    if isinstance(val, datetime):
                        val = val.strftime("%Y-%m-%d %H:%M:%S")
                    else:
                        val = datetime.fromisoformat(str(val)).strftime(
                            "%Y-%m-%d %H:%M:%S"
                        )
                except Exception:
                    pass
            elif key in ("ml_probability", "heuristic_score", "risk_score"):
                if val not in (None, ""):
                    val = f"{float(val) * 100:.1f}%"
                else:
                    val = ""
            elif key == "is_https":
                val = "Yes" if val else "No"
            elif key == "reasons":
                if isinstance(val, list):
                    val = "; ".join(val)
            row[display] = val
        rows.append(row)

    if not rows:
        return pd.DataFrame(columns=list(EXPORT_COLUMNS.values()))

    return pd.DataFrame(rows)


def _generate_csv(df: pd.DataFrame) -> io.BytesIO:
    """Generate a CSV file buffer from a DataFrame."""
    buffer = io.BytesIO()
    df.to_csv(buffer, index=False, encoding="utf-8-sig")
    buffer.seek(0)
    return buffer


def _generate_xlsx(df: pd.DataFrame) -> io.BytesIO:
    """Generate a styled Excel file buffer from a DataFrame."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Security Logs")
        workbook = writer.book
        worksheet = writer.sheets["Security Logs"]

        # Style the header row
        header_fill = PatternFill(
            start_color="0E7490", end_color="0E7490", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=11, name="Consolas")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            bottom=Side(style="thin", color="164E63"),
        )

        for col_idx, col_name in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        # Auto-fit column widths
        for col_idx, col_name in enumerate(df.columns, 1):
            column_letter = worksheet.cell(row=1, column=col_idx).column_letter
            max_length = len(str(col_name))

            for cell in worksheet[column_letter]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            worksheet.column_dimensions[column_letter].width = min(max_length + 4, 50)

        # Style data rows with alternating background
        data_font = Font(size=10, name="Consolas")
        even_fill = PatternFill(
            start_color="F0FDFA", end_color="F0FDFA", fill_type="solid"
        )
        for row_idx in range(2, len(df) + 2):
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                if row_idx % 2 == 0:
                    cell.fill = even_fill

        # Freeze the header row
        worksheet.freeze_panes = "A2"

    buffer.seek(0)
    return buffer


def _generate_pdf(df: pd.DataFrame, start_date: str, end_date: str) -> io.BytesIO:
    """Generate a professional PDF security report using ReportLab."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=20 * mm,
        leftMargin=20 * mm,
        topMargin=20 * mm,
        bottomMargin=20 * mm,
    )

    styles = getSampleStyleSheet()
    story = []

    # Custom styles
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Title"],
        fontSize=22,
        textColor=colors.HexColor("#0E7490"),
        spaceAfter=6,
    )
    subtitle_style = ParagraphStyle(
        "ReportSubtitle",
        parent=styles["Normal"],
        fontSize=10,
        textColor=colors.HexColor("#6B7280"),
        spaceAfter=20,
    )
    section_style = ParagraphStyle(
        "SectionHeader",
        parent=styles["Heading2"],
        fontSize=14,
        textColor=colors.HexColor("#0E7490"),
        spaceBefore=16,
        spaceAfter=8,
    )
    body_style = ParagraphStyle(
        "BodyText",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.HexColor("#1F2937"),
        spaceAfter=4,
    )

    # ── Header ──
    story.append(Paragraph("🛡 Automated API Security Testing Report", title_style))
    story.append(
        Paragraph(
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} &nbsp;|&nbsp; "
            f"Date Range: {start_date} → {end_date}",
            subtitle_style,
        )
    )
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#E5E7EB")))
    story.append(Spacer(1, 10))

    # ── Section 1: Scan Summary ──
    story.append(Paragraph("1. Scan Summary", section_style))
    total = len(df)
    high_count = len(df[df.get("Risk Level", pd.Series()) == "High"]) if "Risk Level" in df.columns else 0
    medium_count = len(df[df.get("Risk Level", pd.Series()) == "Medium"]) if "Risk Level" in df.columns else 0
    low_count = total - high_count - medium_count

    summary_data = [
        ["Metric", "Value"],
        ["Total APIs Scanned", str(total)],
        ["High Risk", str(high_count)],
        ["Medium Risk", str(medium_count)],
        ["Low Risk", str(low_count)],
        ["Scan Period", f"{start_date} to {end_date}"],
    ]
    summary_table = Table(summary_data, colWidths=[160, 200])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0E7490")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0FDFA")]),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 12))

    # ── Section 2: Threat Analysis Table ──
    story.append(Paragraph("2. Threat Analysis", section_style))

    if total > 0:
        # Select key columns for the PDF table
        pdf_cols = ["Domain", "Method", "Risk Score", "Threat Type", "OWASP Category"]
        available_cols = [c for c in pdf_cols if c in df.columns]

        if available_cols:
            table_data = [available_cols]
            for _, row in df.head(30).iterrows():
                table_row = []
                for col in available_cols:
                    val = str(row.get(col, ""))
                    # Truncate long values
                    if len(val) > 35:
                        val = val[:32] + "..."
                    table_row.append(val)
                table_data.append(table_row)

            col_widths = [120, 60, 70, 140, 140][:len(available_cols)]
            threat_table = Table(table_data, colWidths=col_widths)
            threat_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0E7490")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0FDFA")]),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]))
            story.append(threat_table)

            if total > 30:
                story.append(
                    Paragraph(f"<i>Showing top 30 of {total} records</i>", body_style)
                )
    else:
        story.append(Paragraph("No threat data found for the selected date range.", body_style))

    story.append(Spacer(1, 12))

    # ── Section 3: Top Risky APIs ──
    story.append(Paragraph("3. Top Risky APIs", section_style))

    if total > 0 and "Risk Score" in df.columns:
        # Sort by risk score (strip % and convert)
        df_sorted = df.copy()
        df_sorted["_score_num"] = df_sorted["Risk Score"].apply(
            lambda x: float(str(x).replace("%", "")) if x else 0
        )
        df_sorted = df_sorted.sort_values("_score_num", ascending=False).head(5)

        for _, row in df_sorted.iterrows():
            domain = row.get("Domain", "Unknown")
            score = row.get("Risk Score", "N/A")
            threat = row.get("Threat Type", "Unknown")
            owasp = row.get("OWASP Category", "")
            story.append(
                Paragraph(
                    f"<b>{domain}</b> — Risk: {score} | {threat} | {owasp}",
                    body_style,
                )
            )
    else:
        story.append(Paragraph("No high-risk APIs detected.", body_style))

    story.append(Spacer(1, 12))

    # ── Section 4: Recommendations ──
    story.append(Paragraph("4. Security Recommendations", section_style))

    recommendations = [
        "Implement input validation and sanitization on all API endpoints",
        "Enforce authentication and authorization on sensitive endpoints",
        "Encrypt all data in transit using TLS/HTTPS",
        "Use parameterized queries to prevent SQL injection attacks",
        "Implement Content Security Policy (CSP) headers to mitigate XSS",
        "Apply rate limiting to prevent brute-force attacks",
        "Use CORS policies that restrict origins to trusted domains only",
        "Regularly rotate API keys and access tokens",
        "Implement proper error handling — avoid exposing stack traces",
        "Conduct regular security audits and penetration testing",
    ]
    for rec in recommendations:
        story.append(Paragraph(f"• {rec}", body_style))

    # Build PDF
    doc.build(story)
    buffer.seek(0)
    return buffer


@router.get("/export-logs")
async def export_logs(
    start_date: str = Query(..., description="Start date (YYYY-MM-DD)"),
    end_date: str = Query(..., description="End date (YYYY-MM-DD)"),
    format: str = Query("csv", description="Export format: csv, xlsx, or pdf"),
    current_user: dict = Depends(get_current_user),
):
    """
    Export security logs for the authenticated user within a date range.
    Supports CSV, styled Excel (.xlsx), and PDF report formats.
    """
    try:
        logs = await _fetch_logs_for_export(
            user_id=current_user["user_id"],
            start_date=start_date,
            end_date=end_date,
        )

        df = _logs_to_dataframe(logs)

        print(
            f"[export] user={current_user['user_id']} range={start_date}→{end_date} "
            f"format={format} rows={len(df)}",
            flush=True,
        )

        if format == "xlsx":
            buffer = _generate_xlsx(df)
            filename = f"security_logs_{start_date}_to_{end_date}.xlsx"
            media_type = (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        elif format == "pdf":
            buffer = _generate_pdf(df, start_date, end_date)
            filename = f"security_report_{start_date}_to_{end_date}.pdf"
            media_type = "application/pdf"
        else:
            buffer = _generate_csv(df)
            filename = f"security_logs_{start_date}_to_{end_date}.csv"
            media_type = "text/csv"

        return StreamingResponse(
            buffer,
            media_type=media_type,
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
            },
        )

    except HTTPException:
        raise
    except Exception as e:
        print(f"[export] Error: {e}", file=sys.stderr, flush=True)
        raise HTTPException(status_code=500, detail=str(e))
